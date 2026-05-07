# -*- coding: utf-8 -*-
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook(r'C:\Workspace\opencode\douban_rating\豆瓣评分_木三三_20260409.xlsx')
sheet = wb.active

print('工作表:', wb.sheetnames)
print('表头:', [cell.value for cell in sheet[1]])