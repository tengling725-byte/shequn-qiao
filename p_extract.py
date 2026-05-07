import pdfplumber
import pandas as pd
from openpyxl import Workbook

pdf_path = r"C:\my工作\公司_先风广集\04_财务报表\财务报表报送与信息采集（企业会计准则一般企业）-2025-先风广集.pdf"
output_path = r"C:\my工作\公司_先风广集\04_财务报表\财务报表报送与信息采集（企业会计准则一般企业）-2025-先风广集_转换.xlsx"

all_data = {}
current_table = None

with pdfplumber.open(pdf_path) as pdf:
    for i, page in enumerate(pdf.pages):
        tables = page.extract_tables()
        if tables:
            all_data[f"Page_{i+1}"] = tables

wb = Workbook()
ws = wb.active
ws.title = "Extracted"

row = 1
for page_name, tables in all_data.items():
    for table_idx, table in enumerate(tables):
        if table and len(table) > 0:
            ws[f"A{row}"] = f"{page_name}_Table{table_idx+1}"
            row += 1
            for r in table:
                for c_idx, val in enumerate(r):
                    ws.cell(row=row, column=c_idx+1, value=val)
                row += 1
            row += 1

wb.save(output_path)
print(f"Done: {output_path}")